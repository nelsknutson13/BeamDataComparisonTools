"""
Data Merger
-----------
Take a source xlsx (from IbaDataReader or PTWdatareader) with
"Depth Scans" and "Profile Scans"/"Profiles" sheets and distribute
its data into the appropriate destination files in a combined-data folder.

Destination naming convention (already in use):
    {Energy}_{SSD}SSD_PDDData.xlsx        ← depth scans
    {Energy}_{SSD}SSD_ProfileData.xlsx    ← profile scans

Workflow:
  1. Pick source file, destination folder, machine name (sheet name).
  2. Click "Read"  → tool lists every scan in the source and checks each
     against the destination, marking it NEW FILE / NEW SHEET / NEW SCAN
     / EXISTS based on identifier columns only (no Dose comparison).
  3. Toggle checkboxes to deselect scans you don't want to write.
  4. Choose the action for EXISTS scans (Replace or Append).
  5. Click "Write selected".
"""

import os
import tkinter as tk
from tkinter import filedialog, ttk, messagebox, simpledialog

import pandas as pd


DEFAULT_DEST = r"C:\Users\nknutson\OneDrive - Washington University in St. Louis\NGDS QA Consortium\Combined Consortium Data"

DEPTH_SHEET_CANDIDATES   = ['Depth Scans']
PROFILE_SHEET_CANDIDATES = ['Profile Scans', 'Profiles']

CHECKED   = '☑'   # ☑
UNCHECKED = '☐'   # ☐


# ── Helpers ───────────────────────────────────────────────────────────────────

def _classify_table(df):
    """Classify a single flat table: 'profile' if it has an Axis column, else 'depth'."""
    return 'profile' if 'Axis' in df.columns else 'depth'


def _load_source_dfs(source_path, status_cb=None):
    """Return {'depth': df|None, 'profile': df|None} from an xlsx or csv source.

    An xlsx normally holds a depth sheet and/or a profile sheet (named in
    DEPTH/PROFILE_SHEET_CANDIDATES). A csv is a single flat table, classified
    as profile when it has an 'Axis' column and depth otherwise.

    If an xlsx has none of the expected tab names, we don't fail — we warn and
    read whatever data-like sheets are present, classifying each by its columns.
    """
    def _warn(msg):
        if status_cb:
            status_cb(msg)

    result = {'depth': None, 'profile': None}
    ext = os.path.splitext(source_path)[1].lower()

    if ext == '.csv':
        try:
            df = pd.read_csv(source_path)
        except Exception as e:
            raise RuntimeError(f"Cannot open source file: {e}")
        result[_classify_table(df)] = df
        return result

    try:
        xl = pd.ExcelFile(source_path)
    except Exception as e:
        raise RuntimeError(f"Cannot open source file: {e}")

    for kind, candidates in (('depth', DEPTH_SHEET_CANDIDATES),
                             ('profile', PROFILE_SHEET_CANDIDATES)):
        for name in candidates:
            if name in xl.sheet_names:
                result[kind] = pd.read_excel(xl, sheet_name=name)
                break

    # Fallback: none of the expected tabs were found — warn and classify by columns.
    if result['depth'] is None and result['profile'] is None:
        _warn("  WARNING: none of the expected tabs "
              f"{DEPTH_SHEET_CANDIDATES + PROFILE_SHEET_CANDIDATES} were found; "
              "reading sheets by their columns instead.")
        for name in xl.sheet_names:
            df = pd.read_excel(xl, sheet_name=name)
            if 'Energy' not in df.columns or 'SSD' not in df.columns:
                _warn(f"    Skipping tab '{name}' — no Energy/SSD columns.")
                continue
            kind = _classify_table(df)
            if result[kind] is None:
                result[kind] = df
                _warn(f"    Using tab '{name}' as {kind} data.")
            else:
                _warn(f"    Skipping tab '{name}' — already have {kind} data.")

    return result


def _format_ssd(ssd):
    try:
        return f"{int(round(float(ssd)))}"
    except Exception:
        return str(ssd)


def _dest_filename(energy, ssd, kind):
    kind_tag = "PDDData" if kind == 'depth' else "ProfileData"
    return f"{energy}_{_format_ssd(ssd)}SSD_{kind_tag}.xlsx"


def _scan_key_columns(kind, df):
    """Return identifier columns for a scan (filtered to those present in df)."""
    if kind == 'depth':
        cols = ['Energy', 'SSD', 'FS', 'Detector']
    else:
        cols = ['Energy', 'SSD', 'FS', 'Axis', 'Depth', 'Detector']
    return [c for c in cols if c in df.columns]


def _scan_mask(df, key_dict):
    """Boolean mask for rows matching all (col, value) pairs in key_dict."""
    mask = pd.Series([True] * len(df), index=df.index)
    for col, val in key_dict.items():
        if col not in df.columns:
            return pd.Series([False] * len(df), index=df.index)
        if pd.isna(val):
            mask &= df[col].isna()
        else:
            mask &= (df[col] == val)
    return mask


def _check_scan_status(dest_path, sheet_name, key_dict, dest_cache):
    """Return one of NEW FILE / NEW SHEET / NEW SCAN / EXISTS for a single scan.
    dest_cache: {(dest_path, sheet_name): (state, df_or_None)}.
    Loads only the one sheet we need, and only if the file and sheet exist.
    """
    cache_key = (dest_path, sheet_name)
    if cache_key not in dest_cache:
        if not os.path.exists(dest_path):
            dest_cache[cache_key] = ('NO_FILE', None)
        else:
            try:
                xl = pd.ExcelFile(dest_path)
                if sheet_name in xl.sheet_names:
                    dest_cache[cache_key] = ('OK', pd.read_excel(xl, sheet_name=sheet_name))
                else:
                    dest_cache[cache_key] = ('NO_SHEET', None)
            except Exception:
                dest_cache[cache_key] = ('NO_FILE', None)
    state, df = dest_cache[cache_key]
    if state == 'NO_FILE':
        return 'NEW FILE'
    if state == 'NO_SHEET':
        return 'NEW SHEET'
    return 'EXISTS' if _scan_mask(df, key_dict).any() else 'NEW SCAN'


def _ws_to_df(ws):
    """Build a DataFrame from an openpyxl worksheet (first row = header).

    Lets us reuse the workbook we already loaded for writing instead of
    parsing the destination file a second time with pd.read_excel.
    """
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return pd.DataFrame()
    df = pd.DataFrame(rows, columns=list(header))
    return df.dropna(how='all').reset_index(drop=True)


# ── Read: scan source, classify each scan against destination ────────────────

def _coerce_ssd(value):
    """SSD entered as text → number when possible (so it matches stored values)."""
    try:
        f = float(value)
        return int(f) if f.is_integer() else f
    except (TypeError, ValueError):
        return value


def _fill_missing_keys(df, kind, prompt_cb, prompted, status_cb):
    """Ensure df has Energy and SSD. Missing/blank ones are requested via
    prompt_cb and applied to every row. `prompted` caches answers so each
    field is asked once across all source tables. Returns df, or None if the
    user cancels a required prompt.
    """
    for field in ('Energy', 'SSD'):
        missing = field not in df.columns or df[field].isna().all()
        if not missing:
            continue
        if field not in prompted:
            value = prompt_cb(field, kind) if prompt_cb else None
            if value is None or str(value).strip() == '':
                status_cb(f"  WARNING: no '{field}' given for {kind} data — skipping.")
                return None
            prompted[field] = _coerce_ssd(value) if field == 'SSD' else str(value).strip()
        df = df.copy()
        df[field] = prompted[field]
        status_cb(f"  Applied {field}={prompted[field]} to all {kind} scans.")
    return df


def read_scans(source_path, dest_folder, machine_name, status_cb, prompt_cb=None):
    """Return (scans, source_dfs).

    scans: list of dicts with keys {kind, keys, key_cols, dest_file, dest_path, status}
    source_dfs: dict {'depth': df, 'profile': df} for downstream filtering
    prompt_cb: optional (field, kind) -> str|None, used to ask the user for an
        Energy/SSD value when the source doesn't specify one.
    """
    if not os.path.isfile(source_path):
        raise RuntimeError("Source file not found.")
    if not os.path.isdir(dest_folder):
        raise RuntimeError("Destination folder not found.")
    machine_name = machine_name.strip()
    if not machine_name:
        raise RuntimeError("Machine name is required.")

    source_dfs = _load_source_dfs(source_path, status_cb)
    df_depth   = source_dfs['depth']
    df_profile = source_dfs['profile']

    if df_depth is None and df_profile is None:
        raise RuntimeError("Source has no recognised data. Expected an xlsx with "
                           f"{DEPTH_SHEET_CANDIDATES + PROFILE_SHEET_CANDIDATES} "
                           "sheet(s), or a csv with the same columns.")
    scans = []
    dest_cache = {}
    prompted = {}

    for kind, df in (('depth', df_depth), ('profile', df_profile)):
        if df is None or df.empty:
            continue
        df = _fill_missing_keys(df, kind, prompt_cb, prompted, status_cb)
        if df is None:
            continue
        source_dfs[kind] = df   # keep filled-in columns for the write step
        key_cols = _scan_key_columns(kind, df)
        if 'Energy' not in key_cols or 'SSD' not in key_cols:
            continue
        unique = df[key_cols].drop_duplicates().reset_index(drop=True)
        for _, row in unique.iterrows():
            keys = {c: row[c] for c in key_cols}
            energy = str(keys['Energy']).strip()
            ssd    = keys['SSD']
            fname  = _dest_filename(energy, ssd, kind)
            dpath  = os.path.join(dest_folder, fname)
            status = _check_scan_status(dpath, machine_name, keys, dest_cache)
            scans.append({
                'kind': kind,
                'keys': keys,
                'key_cols': key_cols,
                'dest_file': fname,
                'dest_path': dpath,
                'status': status,
            })

    n_new      = sum(1 for s in scans if s['status'].startswith('NEW'))
    n_exist    = sum(1 for s in scans if s['status'] == 'EXISTS')
    n_files    = len({s['dest_file'] for s in scans})
    status_cb(f"Read {len(scans)} scans  ({n_new} NEW, {n_exist} EXISTS) across {n_files} dest file(s).")
    return scans, source_dfs


# ── Write: process selected scans by destination file ────────────────────────

def write_scans(scans, source_dfs, machine_name, conflict_action, status_cb):
    """Write the selected scans into their destination files.

    conflict_action: 'replace' or 'append' (only matters for EXISTS rows).
    """
    if not scans:
        status_cb("No scans selected — nothing to write.")
        return

    # Group selected scans by destination file
    by_dest = {}
    for s in scans:
        by_dest.setdefault(s['dest_path'], []).append(s)

    from openpyxl import load_workbook

    written = replaced = appended = skipped = created = 0
    for dest_path, scan_list in by_dest.items():
        # Load the destination workbook once (used both to read the existing
        # machine sheet and to write back) — avoids parsing the file twice.
        wb = None
        old_df = None
        file_was_new = not os.path.exists(dest_path)
        if not file_was_new:
            try:
                wb = load_workbook(dest_path)
            except Exception as e:
                status_cb(f"  ERROR reading {os.path.basename(dest_path)}: {e}")
                continue
            if machine_name in wb.sheetnames:
                old_df = _ws_to_df(wb[machine_name])

        # If skipping EXISTS scans, drop them from the work list now
        if conflict_action == 'skip' and old_df is not None:
            n_skipped_now = sum(1 for s in scan_list if s['status'] == 'EXISTS')
            scan_list = [s for s in scan_list if s['status'] != 'EXISTS']
            skipped += n_skipped_now
            if not scan_list:
                continue

        # Build the new rows for this file (one DataFrame for the machine sheet)
        new_chunks = []
        for s in scan_list:
            df_src = source_dfs[s['kind']]
            mask = _scan_mask(df_src, s['keys'])
            chunk = df_src[mask].copy()
            if not chunk.empty:
                new_chunks.append(chunk)
        if not new_chunks:
            continue
        new_df = pd.concat(new_chunks, ignore_index=True)

        # Merge with the existing machine sheet (if any)
        if old_df is not None:
            if conflict_action == 'replace':
                drop_mask = pd.Series([False] * len(old_df), index=old_df.index)
                for s in scan_list:
                    if s['status'] == 'EXISTS':
                        drop_mask |= _scan_mask(old_df, s['keys'])
                kept = old_df[~drop_mask]
                final_df = pd.concat([kept, new_df], ignore_index=True)
                replaced += sum(1 for s in scan_list if s['status'] == 'EXISTS')
                written  += sum(1 for s in scan_list if s['status'] != 'EXISTS')
            else:  # append (skip already filtered above, so no EXISTS rows here for skip)
                final_df = pd.concat([old_df, new_df], ignore_index=True)
                appended += sum(1 for s in scan_list if s['status'] == 'EXISTS')
                written  += sum(1 for s in scan_list if s['status'] != 'EXISTS')
        else:
            final_df = new_df
            written += len(scan_list)

        # Write only the target sheet — preserve other sheets in place
        try:
            if file_was_new:
                with pd.ExcelWriter(dest_path, engine='openpyxl') as w:
                    final_df.to_excel(w, sheet_name=machine_name, index=False)
                created += 1
            else:
                if machine_name in wb.sheetnames:
                    del wb[machine_name]
                ws = wb.create_sheet(machine_name)
                ws.append(list(final_df.columns))
                for row in final_df.itertuples(index=False, name=None):
                    ws.append(row)
                wb.save(dest_path)
        except Exception as e:
            status_cb(f"  ERROR writing {os.path.basename(dest_path)}: {e}")
            continue

        status_cb(f"  Wrote {len(new_df)} rows to '{machine_name}' in "
                  f"{os.path.basename(dest_path)} ({len(scan_list)} scan(s))")

    status_cb("")
    status_cb("=" * 60)
    status_cb(f"  Scans written: {written}   Replaced: {replaced}   "
              f"Appended: {appended}   Skipped: {skipped}   New files: {created}")
    status_cb("=" * 60)


# ── GUI ──────────────────────────────────────────────────────────────────────

root = tk.Tk()
root.title("Data Merger v0.2")

source_var   = tk.StringVar(master=root)
dest_var     = tk.StringVar(master=root, value=DEFAULT_DEST)
machine_var  = tk.StringVar(master=root)
conflict_var = tk.StringVar(master=root, value='replace')   # replace | append

_scans      = []   # list of scan info dicts (mirrors tree rows)
_source_dfs = {}   # {'depth': df, 'profile': df}

main = ttk.Frame(root, padding=10)
main.grid(row=0, column=0, sticky="nsew")


def _log(msg):
    output.insert(tk.END, str(msg) + "\n")
    output.see(tk.END)
    output.update_idletasks()


def choose_source():
    p = filedialog.askopenfilename(filetypes=[("Excel or CSV", "*.xlsx *.xls *.csv"),
                                              ("Excel", "*.xlsx *.xls"),
                                              ("CSV", "*.csv"),
                                              ("All", "*.*")])
    if p:
        source_var.set(p)


def choose_dest():
    d = filedialog.askdirectory(initialdir=dest_var.get() or DEFAULT_DEST)
    if d:
        dest_var.set(d)


def _format_scan_row(scan):
    k = scan['keys']
    def _v(col):
        if col not in k:
            return ''
        v = k[col]
        if isinstance(v, float) and v.is_integer():
            return f"{int(v)}"
        return str(v) if not pd.isna(v) else ''
    return (CHECKED, scan['kind'],
            _v('Energy'), _v('SSD'), _v('FS'),
            _v('Axis'), _v('Depth'), _v('Detector'),
            scan['dest_file'], scan['status'])


def do_read():
    global _scans, _source_dfs
    output.delete("1.0", tk.END)
    tree.delete(*tree.get_children())
    _scans = []
    _source_dfs = {}
    def _prompt_value(field, kind):
        return simpledialog.askstring(
            f"{field} required",
            f"The {kind} data has no '{field}'.\n"
            f"Enter the {field} to apply to all {kind} scans:",
            parent=root)

    try:
        scans, source_dfs = read_scans(source_var.get(), dest_var.get(),
                                       machine_var.get(), _log, _prompt_value)
    except Exception as e:
        _log(f"\n[ERROR] {e}")
        messagebox.showerror("Error", str(e))
        return
    _scans = scans
    _source_dfs = source_dfs
    for s in _scans:
        tree.insert('', 'end', values=_format_scan_row(s))


def _selected_scans():
    selected = []
    for iid, scan in zip(tree.get_children(), _scans):
        if tree.set(iid, 'check') == CHECKED:
            selected.append(scan)
    return selected


def do_write():
    if not _scans:
        messagebox.showinfo("Nothing to write", "Click Read first to load scans.")
        return
    selected = _selected_scans()
    if not selected:
        messagebox.showinfo("Nothing selected", "Select at least one scan.")
        return
    try:
        write_scans(selected, _source_dfs, machine_var.get().strip(),
                    conflict_var.get(), _log)
    except Exception as e:
        _log(f"\n[ERROR] {e}")
        messagebox.showerror("Error", str(e))


def _toggle_row(iid):
    tree.set(iid, 'check', UNCHECKED if tree.set(iid, 'check') == CHECKED else CHECKED)


def _on_tree_click(event):
    region = tree.identify_region(event.x, event.y)
    if region != 'cell':
        return
    if tree.identify_column(event.x) != '#1':
        return
    iid = tree.identify_row(event.y)
    if iid:
        _toggle_row(iid)


def select_all():
    for iid in tree.get_children():
        tree.set(iid, 'check', CHECKED)


def deselect_all():
    for iid in tree.get_children():
        tree.set(iid, 'check', UNCHECKED)


# Row 0: source
ttk.Label(main, text="Source file:").grid(row=0, column=0, sticky="w")
ttk.Entry(main, textvariable=source_var, width=70).grid(row=0, column=1, padx=5)
ttk.Button(main, text="Browse…", command=choose_source).grid(row=0, column=2)

# Row 1: destination
ttk.Label(main, text="Destination folder:").grid(row=1, column=0, sticky="w", pady=(6, 0))
ttk.Entry(main, textvariable=dest_var, width=70).grid(row=1, column=1, padx=5, pady=(6, 0))
ttk.Button(main, text="Browse…", command=choose_dest).grid(row=1, column=2, pady=(6, 0))

# Row 2: machine name
ttk.Label(main, text="Machine name (sheet):").grid(row=2, column=0, sticky="w", pady=(6, 0))
ttk.Entry(main, textvariable=machine_var, width=30).grid(row=2, column=1, padx=5, sticky="w", pady=(6, 0))

# Row 3: conflict action
ttk.Label(main, text="On EXISTS conflict:").grid(row=3, column=0, sticky="w", pady=(6, 0))
conflict_frame = ttk.Frame(main)
conflict_frame.grid(row=3, column=1, sticky="w", padx=5, pady=(6, 0))
ttk.Radiobutton(conflict_frame, text="Replace",
                variable=conflict_var, value='replace').pack(side='left', padx=2)
ttk.Radiobutton(conflict_frame, text="Append",
                variable=conflict_var, value='append').pack(side='left', padx=8)
ttk.Radiobutton(conflict_frame, text="Skip",
                variable=conflict_var, value='skip').pack(side='left', padx=2)

# Row 4: action buttons
btn_frame = ttk.Frame(main)
btn_frame.grid(row=4, column=0, columnspan=3, pady=10, sticky="w")
ttk.Button(btn_frame, text="Read",            command=do_read     ).pack(side='left', padx=4)
ttk.Button(btn_frame, text="Select all",      command=select_all  ).pack(side='left', padx=12)
ttk.Button(btn_frame, text="Deselect all",    command=deselect_all).pack(side='left', padx=4)
ttk.Button(btn_frame, text="Write selected",  command=do_write    ).pack(side='left', padx=24)

# Row 5: tree of scans
tree_frame = ttk.Frame(main)
tree_frame.grid(row=5, column=0, columnspan=3, sticky="nsew")
columns = ('check', 'kind', 'Energy', 'SSD', 'FS', 'Axis', 'Depth', 'Detector', 'Dest File', 'Status')
tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=18, selectmode='none')
widths = {'check': 30, 'kind': 60, 'Energy': 60, 'SSD': 50, 'FS': 50,
          'Axis': 50, 'Depth': 60, 'Detector': 90, 'Dest File': 240, 'Status': 90}
for c in columns:
    tree.heading(c, text=c.title() if c != 'check' else '')
    tree.column(c, width=widths.get(c, 80), anchor='w', stretch=False)
tree.grid(row=0, column=0, sticky="nsew")
vsb = ttk.Scrollbar(tree_frame, orient="vertical",   command=tree.yview)
hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
vsb.grid(row=0, column=1, sticky="ns")
hsb.grid(row=1, column=0, sticky="ew")
tree.bind('<Button-1>', _on_tree_click)

# Row 6: log
output_frame = ttk.Frame(main)
output_frame.grid(row=6, column=0, columnspan=3, sticky="w", pady=(10, 0))
output = tk.Text(output_frame, height=8, width=110, wrap='none', font=('Courier', 9))
output.grid(row=0, column=0, sticky="nsew")
ovsb = ttk.Scrollbar(output_frame, orient="vertical", command=output.yview)
ovsb.grid(row=0, column=1, sticky="ns")
output.configure(yscrollcommand=ovsb.set)


root.mainloop()
