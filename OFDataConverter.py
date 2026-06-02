"""
Output Factor Data Converter
----------------------------
Convert a wide-format 2D output factor table (Y rows × X columns) into the
combined long-format dataset used by OFCompare.

Long-format columns:
    Energy, SSD, Depth, SN, Detector, FS_X, FS_Y, Scp

Workflow:
  1. Pick a source file + sheet.
  2. Fill in metadata (Energy, SSD, Depth, SN, Detector).
  3. Tool auto-detects whether the data is already normalized (cell at the
     reference FS ≈ 1.0). If not, it normalizes by dividing every cell by
     the reference cell.
  4. Click "Read & Preview" to see what would be added.
  5. Click "Write" to append/replace into the combined xlsx.

Conflict policy is per-key (Energy, SSD, Depth, SN, Detector, FS_X, FS_Y):
  - Skip:    leave existing row, drop new row
  - Replace: overwrite existing row with new value
"""

import os
import re
import tkinter as tk
from tkinter import filedialog, ttk, messagebox

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


DEFAULT_DEST = (r"C:\Users\nknutson\OneDrive - Washington University in St. Louis"
                r"\NGDS QA Consortium\Combined Consortium Data"
                r"\90 SSD\OF_Data\NGDSOutputFactorData.xlsx")
DEST_SHEET   = 'OutputFactors'

KEY_COLS  = ['Energy', 'SSD', 'Depth', 'SN', 'Detector', 'FS_X', 'FS_Y']
LONG_COLS = KEY_COLS + ['Scp']

ENERGY_TOKENS = ['6X', '10X', '15X', '6FFF', '8FFF', '10FFF']
NORMALIZED_TOL = 1e-3   # how close to 1.0 the ref cell must be to skip renormalization


# ── Wide → long conversion ───────────────────────────────────────────────────

def _to_float(v):
    """Try to coerce to float, tolerating comma decimals ('0,762') and stray whitespace.
    Returns None on failure.
    """
    if v is None:
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        if isinstance(v, str):
            s = v.strip().replace(',', '.')
            try:
                f = float(s)
            except ValueError:
                return None
        else:
            return None
    return f if f == f else None   # excludes NaN


def _is_numeric(v):
    return _to_float(v) is not None


def wide_to_long(df_raw):
    """Convert a wide 2D (Y rows × X columns) sheet into a long DataFrame
    with FS_X, FS_Y, Scp columns. Auto-detects header row/column.
    Tolerates label cells in the top-left corner and comma decimal separators.
    """
    arr = df_raw.values
    n_rows, n_cols = arr.shape

    # Find the X header row + the column where numeric X values start.
    # Some pastes have label cells like "6X" / "X" in the top-left corner,
    # so we scan for the first row whose tail (col 1..N) is mostly numeric
    # and remember where the numeric run starts.
    x_header_row = None
    x_start_col  = 1
    for r in range(min(n_rows, 6)):
        if n_cols < 2:
            continue
        for start in range(1, n_cols):
            tail = arr[r, start:]
            if len(tail) < 3:
                continue
            numeric_count = sum(_is_numeric(v) for v in tail)
            if numeric_count >= max(3, int(0.6 * len(tail))):
                x_header_row = r
                x_start_col  = start
                break
        if x_header_row is not None:
            break
    if x_header_row is None:
        raise RuntimeError("Could not find a numeric X-axis header row.")

    x_vals, x_cols = [], []
    for c in range(x_start_col, n_cols):
        v = _to_float(arr[x_header_row, c])
        if v is not None:
            x_vals.append(v)
            x_cols.append(c)
    if not x_vals:
        raise RuntimeError("No numeric X values found.")

    # Find the Y column: first column at/before x_start_col with mostly numeric values
    # in rows after x_header_row.
    y_col = None
    for c in range(min(x_start_col, n_cols)):
        col_vals = arr[x_header_row + 1:, c]
        if len(col_vals) == 0:
            continue
        if sum(_is_numeric(v) for v in col_vals) >= max(3, int(0.6 * len(col_vals))):
            y_col = c
            break
    if y_col is None:
        y_col = 0   # fall back to leftmost column

    rows = []
    for r in range(x_header_row + 1, n_rows):
        y = _to_float(arr[r, y_col])
        if y is None:
            continue
        for c, x in zip(x_cols, x_vals):
            v = _to_float(arr[r, c])
            if v is not None:
                rows.append({'FS_X': x, 'FS_Y': y, 'Scp': v})
    if not rows:
        raise RuntimeError("No data rows below the X header.")
    return pd.DataFrame(rows)


def _ref_value(df, ref_x, ref_y):
    """Return the Scp at (ref_x, ref_y), or None if not present."""
    m = (np.isclose(df['FS_X'], ref_x)) & (np.isclose(df['FS_Y'], ref_y))
    if not m.any():
        return None
    return float(df.loc[m, 'Scp'].iloc[0])


def normalize(df, ref_x, ref_y, force=False):
    """If the value at (ref_x, ref_y) isn't already 1.0 (within tol) or if
    force=True, divide all Scp by that value. Returns (df, was_normalized).
    """
    ref = _ref_value(df, ref_x, ref_y)
    if ref is None:
        raise RuntimeError(f"Reference FS ({ref_x}×{ref_y}) not found in data.")
    if not force and abs(ref - 1.0) <= NORMALIZED_TOL:
        return df.copy(), False
    out = df.copy()
    out['Scp'] = out['Scp'] / ref
    return out, True


def _guess_energy(text):
    if not text:
        return None
    s = str(text).upper().replace(' ', '').replace('_', '')
    for tok in ENERGY_TOKENS:
        if tok in s:
            return tok
    return None


def _guess_sn(text):
    """Pull a serial-number-like token (e.g. 'SN 21', 'SN21', 'SN_1003') from a path or string."""
    if not text:
        return None
    m = re.search(r'\bSN[\s_-]*\d+\b', str(text), re.IGNORECASE)
    if not m:
        return None
    raw = m.group(0).upper()
    # Normalize to "SN <digits>"
    digits = re.search(r'\d+', raw).group(0)
    return f"SN {digits}"


def _guess_ssd(text):
    if not text:
        return None
    s = str(text).upper()
    m = re.search(r'(\d{2,3})\s*SSD|SSD\s*(\d{2,3})', s)
    if m:
        return int(m.group(1) or m.group(2))
    return None


# ── Combined xlsx I/O ────────────────────────────────────────────────────────

def read_combined(dest_path):
    """Return the combined long-format DataFrame, or empty df with the right columns."""
    if not os.path.exists(dest_path):
        return pd.DataFrame(columns=LONG_COLS)
    try:
        xl = pd.ExcelFile(dest_path)
        if DEST_SHEET in xl.sheet_names:
            df = pd.read_excel(xl, sheet_name=DEST_SHEET)
        else:
            df = pd.DataFrame(columns=LONG_COLS)
    except Exception as e:
        raise RuntimeError(f"Cannot read combined file: {e}")
    for c in LONG_COLS:
        if c not in df.columns:
            df[c] = pd.NA
    return df[LONG_COLS]


def write_combined(dest_path, df):
    """Write df to DEST_SHEET in dest_path, preserving any other sheets."""
    os.makedirs(os.path.dirname(dest_path), exist_ok=True)
    if not os.path.exists(dest_path):
        with pd.ExcelWriter(dest_path, engine='openpyxl') as w:
            df.to_excel(w, sheet_name=DEST_SHEET, index=False)
        return
    wb = load_workbook(dest_path)
    if DEST_SHEET in wb.sheetnames:
        del wb[DEST_SHEET]
    ws = wb.create_sheet(DEST_SHEET)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    wb.save(dest_path)


def merge_long(existing, new, conflict_action):
    """Merge new long-format rows into existing. Returns (merged_df, n_added,
    n_skipped, n_replaced). Dedup key = KEY_COLS.
    """
    if existing.empty:
        return new.copy(), len(new), 0, 0

    # Round numeric keys lightly to avoid floating-point near-misses
    def _norm(df):
        out = df.copy()
        for c in ('FS_X', 'FS_Y', 'SSD', 'Depth'):
            if c in out.columns:
                out[c] = pd.to_numeric(out[c], errors='coerce').round(3)
        for c in ('Energy', 'SN', 'Detector'):
            if c in out.columns:
                out[c] = out[c].astype(str).str.strip()
        return out

    a = _norm(existing).reset_index(drop=True)
    b = _norm(new).reset_index(drop=True)

    a_keys = a[KEY_COLS].apply(tuple, axis=1)
    b_keys = b[KEY_COLS].apply(tuple, axis=1)

    overlap = b_keys.isin(set(a_keys))
    n_overlap = int(overlap.sum())

    if conflict_action == 'skip':
        b_to_add = b[~overlap]
        merged = pd.concat([a, b_to_add], ignore_index=True)
        return merged, len(b_to_add), n_overlap, 0
    else:  # replace
        # drop overlapping keys from a, keep all of b
        a_kept = a[~a_keys.isin(set(b_keys[overlap]))]
        merged = pd.concat([a_kept, b], ignore_index=True)
        return merged, len(b) - n_overlap, 0, n_overlap


# ── GUI ──────────────────────────────────────────────────────────────────────

root = tk.Tk()
root.title("Output Factor Data Converter v0.1")

src_path = tk.StringVar(master=root)
src_sheet = tk.StringVar(master=root)
dest_path = tk.StringVar(master=root, value=DEFAULT_DEST)

energy_var   = tk.StringVar(master=root)
ssd_var      = tk.StringVar(master=root, value='90')
depth_var    = tk.StringVar(master=root, value='10')
sn_var       = tk.StringVar(master=root)
detector_var = tk.StringVar(master=root)

ref_x_var    = tk.StringVar(master=root, value='10.5')
ref_y_var    = tk.StringVar(master=root, value='10.5')
force_norm   = tk.BooleanVar(master=root, value=False)

conflict_var = tk.StringVar(master=root, value='replace')

_preview_df = None

main = ttk.Frame(root, padding=10)
main.grid(row=0, column=0, sticky="nsew")


def _log(msg):
    output.insert(tk.END, str(msg) + "\n")
    output.see(tk.END)
    output.update_idletasks()


def choose_src():
    p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
    if not p:
        return
    src_path.set(p)
    try:
        sheets = pd.ExcelFile(p).sheet_names
    except Exception as e:
        messagebox.showerror("Error", f"Cannot open file: {e}")
        return
    sheet_combo['values'] = sheets
    if sheets:
        src_sheet.set(sheets[0])
        _on_sheet_change()
    # Auto-fill from filename + directory path
    g_ssd = _guess_ssd(p)
    if g_ssd is not None and not ssd_var.get():
        ssd_var.set(str(g_ssd))
    g_energy = _guess_energy(p)
    if g_energy and not energy_var.get():
        energy_var.set(g_energy)
    g_sn = _guess_sn(p)
    if g_sn and not sn_var.get():
        sn_var.set(g_sn)


def _on_sheet_change(*_):
    sheet = src_sheet.get()
    path  = src_path.get()
    # Try sheet name first (more specific), fall back to full path.
    for source in (sheet, path):
        g_e = _guess_energy(source)
        if g_e and not energy_var.get():
            energy_var.set(g_e)
        g_ssd = _guess_ssd(source)
        if g_ssd is not None and not ssd_var.get():
            ssd_var.set(str(g_ssd))
        g_sn = _guess_sn(source)
        if g_sn and not sn_var.get():
            sn_var.set(g_sn)


def choose_dest():
    p = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                     filetypes=[("Excel", "*.xlsx")],
                                     initialfile=os.path.basename(dest_path.get() or DEFAULT_DEST))
    if p:
        dest_path.set(p)


def _preview_from_raw(raw, source_tag=''):
    """Run the full preview/normalize/metadata pipeline on a raw wide DataFrame."""
    global _preview_df
    output.delete("1.0", tk.END)
    if source_tag:
        _log(f"Source: {source_tag}")
    try:
        df = wide_to_long(raw)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to parse: {e}")
        return

    try:
        ref_x = float(ref_x_var.get()); ref_y = float(ref_y_var.get())
    except ValueError:
        messagebox.showerror("Error", "Reference X / Y must be numeric.")
        return

    try:
        df, was_norm = normalize(df, ref_x, ref_y, force=force_norm.get())
    except Exception as e:
        messagebox.showerror("Error", f"Normalization failed: {e}")
        return

    if was_norm:
        _log(f"Normalized: divided all Scp by value at ({ref_x}, {ref_y}).")
    else:
        _log(f"Already normalized at ({ref_x}, {ref_y}) — no change.")

    energy = energy_var.get().strip()
    sn     = sn_var.get().strip()
    det    = detector_var.get().strip()
    try:
        ssd   = float(ssd_var.get())
        depth = float(depth_var.get())
    except ValueError:
        messagebox.showerror("Error", "SSD and Depth must be numeric.")
        return
    missing = [name for name, val in (('Energy', energy), ('SN', sn), ('Detector', det)) if not val]
    if missing:
        messagebox.showerror("Error", f"Missing metadata: {', '.join(missing)}")
        return

    df['Energy']   = energy
    df['SSD']      = ssd
    df['Depth']    = depth
    df['SN']       = sn
    df['Detector'] = det
    df = df[LONG_COLS]
    _preview_df = df

    _log(f"Parsed {len(df)} rows  ({df['FS_X'].nunique()} X × {df['FS_Y'].nunique()} Y).")
    _log(f"  Energy={energy}  SSD={ssd:g}  Depth={depth:g}  SN={sn}  Detector={det}")

    dp = dest_path.get().strip()
    if dp:
        try:
            existing = read_combined(dp)
        except Exception as e:
            _log(f"  WARNING: cannot read destination ({e})")
            return
        if existing.empty:
            _log(f"  Destination is empty / new — all {len(df)} rows would be added.")
        else:
            def _norm(d):
                out = d.copy()
                for c in ('FS_X', 'FS_Y', 'SSD', 'Depth'):
                    if c in out.columns:
                        out[c] = pd.to_numeric(out[c], errors='coerce').round(3)
                for c in ('Energy', 'SN', 'Detector'):
                    if c in out.columns:
                        out[c] = out[c].astype(str).str.strip()
                return out
            a_keys = set(_norm(existing)[KEY_COLS].apply(tuple, axis=1))
            b_keys = _norm(df)[KEY_COLS].apply(tuple, axis=1)
            n_over = int(b_keys.isin(a_keys).sum())
            _log(f"  Destination has {len(existing)} rows already.")
            _log(f"  Of {len(df)} new rows, {n_over} match existing keys (will be {conflict_var.get()}d).")


def do_preview():
    p, sh = src_path.get().strip(), src_sheet.get().strip()
    if not p or not sh:
        messagebox.showerror("Error", "Pick a source file and sheet.")
        return
    try:
        raw = pd.read_excel(p, sheet_name=sh, header=None)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to read sheet: {e}")
        return
    _preview_from_raw(raw, source_tag=f"{os.path.basename(p)} :: {sh}")


def do_paste():
    """Open a popup with a text area, parse the pasted tab/space-delimited
    wide-format table, and feed it through the same preview pipeline.
    """
    dlg = tk.Toplevel(root)
    dlg.title("Paste OF data (tab- or space-delimited wide-format)")
    dlg.transient(root)
    info = ("Paste a 2D table here — first row is X values (with optional 'x' header in cell A1),\n"
            "first column is Y values (with optional 'y' label), interior cells are Scp readings.\n"
            "Tab-separated (e.g. straight from Excel) or comma/space-separated all work.")
    ttk.Label(dlg, text=info, padding=8, justify='left').pack(anchor='w')
    txt_frame = ttk.Frame(dlg, padding=(8, 0, 8, 0))
    txt_frame.pack(fill='both', expand=True)
    txt = tk.Text(txt_frame, wrap='none', font=('Courier', 9), width=110, height=20)
    txt.grid(row=0, column=0, sticky='nsew')
    vsb = ttk.Scrollbar(txt_frame, orient='vertical',   command=txt.yview)
    hsb = ttk.Scrollbar(txt_frame, orient='horizontal', command=txt.xview)
    txt.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
    vsb.grid(row=0, column=1, sticky='ns')
    hsb.grid(row=1, column=0, sticky='ew')
    txt_frame.grid_rowconfigure(0, weight=1); txt_frame.grid_columnconfigure(0, weight=1)
    txt.focus_set()

    btns = ttk.Frame(dlg, padding=8)
    btns.pack(fill='x')

    def _parse_and_preview():
        s = txt.get('1.0', tk.END)
        if not s.strip():
            messagebox.showerror("Empty", "Nothing to parse.", parent=dlg)
            return
        import io as _io
        # Try whitespace-tolerant parse: tabs OR ≥2 spaces OR commas
        raw = None
        for sep in ('\t', r'\s{2,}', ','):
            try:
                cand = pd.read_csv(_io.StringIO(s), sep=sep, header=None, engine='python')
                if cand.shape[1] >= 2:
                    raw = cand
                    break
            except Exception:
                continue
        if raw is None:
            messagebox.showerror("Parse error",
                                 "Could not parse the pasted text as a table.", parent=dlg)
            return
        dlg.destroy()
        _preview_from_raw(raw, source_tag='(pasted data)')

    ttk.Button(btns, text="Parse & Preview", command=_parse_and_preview).pack(side='right', padx=4)
    ttk.Button(btns, text="Cancel",          command=dlg.destroy        ).pack(side='right', padx=4)


def do_write():
    if _preview_df is None:
        messagebox.showinfo("Preview first", "Click 'Read & Preview' first.")
        return
    dp = dest_path.get().strip()
    if not dp:
        messagebox.showerror("Error", "Pick a destination file.")
        return
    try:
        existing = read_combined(dp)
    except Exception as e:
        messagebox.showerror("Error", str(e))
        return
    merged, n_added, n_skipped, n_replaced = merge_long(existing, _preview_df, conflict_var.get())
    try:
        write_combined(dp, merged)
    except Exception as e:
        messagebox.showerror("Error writing", str(e))
        return
    _log(f"\nWrote {len(merged)} total rows to {os.path.basename(dp)}")
    _log(f"  Added: {n_added}   Skipped: {n_skipped}   Replaced: {n_replaced}")


# Layout
ttk.Label(main, text="Source file:").grid(row=0, column=0, sticky="w")
ttk.Entry(main, textvariable=src_path, width=70).grid(row=0, column=1, padx=4)
ttk.Button(main, text="Browse…", command=choose_src).grid(row=0, column=2)

ttk.Label(main, text="Source sheet:").grid(row=1, column=0, sticky="w", pady=(4, 0))
sheet_combo = ttk.Combobox(main, textvariable=src_sheet, width=20, state='readonly')
sheet_combo.grid(row=1, column=1, sticky="w", padx=4, pady=(4, 0))
src_sheet.trace_add('write', _on_sheet_change)

# Metadata frame
meta_frame = ttk.LabelFrame(main, text="Metadata", padding=6)
meta_frame.grid(row=2, column=0, columnspan=3, sticky="we", pady=(8, 0))
ttk.Label(meta_frame, text="Energy:").grid(row=0, column=0, sticky="w")
ttk.Entry(meta_frame, textvariable=energy_var, width=10).grid(row=0, column=1, padx=4)
ttk.Label(meta_frame, text="SSD [cm]:").grid(row=0, column=2, sticky="w", padx=(12, 0))
ttk.Entry(meta_frame, textvariable=ssd_var, width=8).grid(row=0, column=3, padx=4)
ttk.Label(meta_frame, text="Depth [cm]:").grid(row=0, column=4, sticky="w", padx=(12, 0))
ttk.Entry(meta_frame, textvariable=depth_var, width=8).grid(row=0, column=5, padx=4)
ttk.Label(meta_frame, text="SN:").grid(row=1, column=0, sticky="w", pady=(4, 0))
ttk.Entry(meta_frame, textvariable=sn_var, width=20).grid(row=1, column=1, columnspan=2, padx=4, sticky="w", pady=(4, 0))
ttk.Label(meta_frame, text="Detector:").grid(row=1, column=3, sticky="w", padx=(12, 0), pady=(4, 0))
ttk.Entry(meta_frame, textvariable=detector_var, width=20).grid(row=1, column=4, columnspan=2, padx=4, sticky="w", pady=(4, 0))

# Normalization frame
norm_frame = ttk.LabelFrame(main, text="Normalization", padding=6)
norm_frame.grid(row=3, column=0, columnspan=3, sticky="we", pady=(8, 0))
ttk.Label(norm_frame, text="Reference FS  X:").grid(row=0, column=0, sticky="w")
ttk.Entry(norm_frame, textvariable=ref_x_var, width=8).grid(row=0, column=1, padx=4)
ttk.Label(norm_frame, text="Y:").grid(row=0, column=2)
ttk.Entry(norm_frame, textvariable=ref_y_var, width=8).grid(row=0, column=3, padx=4)
ttk.Checkbutton(norm_frame, text="Force renormalize (override auto-detect)",
                variable=force_norm).grid(row=0, column=4, padx=(12, 0))

# Destination + conflict
ttk.Label(main, text="Destination file:").grid(row=4, column=0, sticky="w", pady=(8, 0))
ttk.Entry(main, textvariable=dest_path, width=70).grid(row=4, column=1, padx=4, pady=(8, 0))
ttk.Button(main, text="Browse…", command=choose_dest).grid(row=4, column=2, pady=(8, 0))

conflict_frame = ttk.Frame(main)
conflict_frame.grid(row=5, column=0, columnspan=3, sticky="w", pady=(4, 0))
ttk.Label(conflict_frame, text="On conflict:").pack(side='left')
ttk.Radiobutton(conflict_frame, text="Replace", variable=conflict_var, value='replace').pack(side='left', padx=4)
ttk.Radiobutton(conflict_frame, text="Skip",    variable=conflict_var, value='skip').pack(side='left', padx=4)

# Action buttons
btns = ttk.Frame(main)
btns.grid(row=6, column=0, columnspan=3, sticky="w", pady=(8, 0))
ttk.Button(btns, text="Read & Preview", command=do_preview).pack(side='left', padx=4)
ttk.Button(btns, text="Paste data…",    command=do_paste  ).pack(side='left', padx=4)
ttk.Button(btns, text="Write",          command=do_write  ).pack(side='left', padx=12)

# Log
output = tk.Text(main, height=14, width=110, font=('Courier', 9), wrap='none')
output.grid(row=7, column=0, columnspan=3, sticky="nsew", pady=(8, 0))


root.mainloop()
